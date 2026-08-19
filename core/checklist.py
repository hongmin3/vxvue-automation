# -*- coding: utf-8 -*-
r"""체크리스트 xlsx에 TC별 자동화 판정 결과를 기록한다.

Bellalun `auto/core/checklist.py`를 이식하되 VXvue 체크리스트의 실제 구조에
맞췄다(2026-08-19 실측).

## 대상 문서 구조 (실측)

```
Windows Update 호환성 검증 Checklist_VXvue_R-25-774.xlsx
  시트 "Summary"    문서번호(C2=R-25-774), 대상 버전 목록
  시트 "Checklist"  <- 결과를 기록하는 시트
     1~5행  A열=항목명 / K열=값   OS, OS Version, OS Build Version,
                                  Viewer Version, VX.LIVE.SERVER
     6행    헤더:  A=STC Category  B=TC ID  C=Function  D=Func_01
                   E=Func_02  F=Title  G=Precondition  H=Step Description
                   I=Expected Result  J=Test Data  K=Result  L=Comment
     7행~   TC_WindowsUpdate_01 ~ 15
```

## 원본을 건드리지 않는다

`VXvue/CLAUDE.md` 3절: "기존 TC(엑셀)의 이력·TC ID·Pass/Fail/Issue 열은 이유
없이 삭제·재작성하지 않는다." 그래서

- 원본 파일은 **읽기만** 하고, `shutil.copyfile`로 뜬 사본에 기록한다.
- 사람이 손으로 채운 `K열 Result` / `L열 Comment`는 **그대로 둔다.** 자동화
  판정은 그 오른쪽에 새 열(`자동화 판정`, `판정 일시`, ...)로만 덧붙인다.
- 1~5행의 환경 값(K열)도 덮어쓰지 않는다. 자동화가 실측한 값은 같은 행의
  **새 열**에 나란히 적어 사람이 차이를 볼 수 있게 한다.

## 기록하지 않는 것

자동화가 수행하지 않은 TC는 `미수행`으로 남긴다. 실행하지 않은 것을 빈칸으로
두면 "확인했는데 이상 없음"으로 오해되므로, 회귀 러너와 같은 원칙
(`core/regression.py`)을 적용해 **수행/미수행을 구분해서** 적는다.
"""

import os
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .result import BLOCKED, FAIL, MANUAL, PASS, SKIP

CHECKLIST_NAME = "Windows Update 호환성 검증 Checklist_VXvue_R-25-774.xlsx"
CHECKLIST_SHEET = "Checklist"
TC_COL_HEADER = "TC ID"

# 1~5행에서 자동화 실측값을 나란히 적을 항목.
# A열 라벨 -> (collect_env의 하위 딕셔너리, 그 안의 키).
# `core/result.collect_env()`는 {"windows": {...}, "packages": {...}} 구조를
# 돌려준다 — 실제 키 이름을 그대로 쓴다(추측하지 않는다).
ENV_ROW_LABELS = {
    "OS":               ("windows", "OS"),
    "OS Version":       ("windows", "OS Version"),
    "OS Build Version": ("windows", "OS Build Version"),
    "Viewer Version":   ("packages", "VXvue"),
    "VX.LIVE.SERVER":   ("packages", "VX.LIVE.SERVER"),
}

RESULT_HEADERS = ["자동화 판정", "판정 일시", "PASS", "FAIL", "MANUAL", "SKIP",
                  "BLOCKED", "실패 항목", "수동 확인 항목", "증적"]

NOT_RUN = "미수행"

FILLS = {
    PASS:    PatternFill("solid", fgColor="C6EFCE"),
    FAIL:    PatternFill("solid", fgColor="FFC7CE"),
    MANUAL:  PatternFill("solid", fgColor="FFEB9C"),
    SKIP:    PatternFill("solid", fgColor="E7E6E6"),
    BLOCKED: PatternFill("solid", fgColor="D9D2E9"),
    NOT_RUN: PatternFill("solid", fgColor="F2F2F2"),
}
FONTS = {
    PASS:    Font(color="006100", bold=True),
    FAIL:    Font(color="9C0006", bold=True),
    MANUAL:  Font(color="9C6500", bold=True),
    SKIP:    Font(color="808080"),
    BLOCKED: Font(color="5B2D8E", bold=True),
    NOT_RUN: Font(color="A6A6A6"),
}


def source_path(cfg, root=None):
    """체크리스트 원본을 PC 독립적으로 찾는다.

    `config.json > checklist_xlsx`가 있으면 최우선으로 쓰되 **실제로 존재할
    때만** 쓴다. Bellalun에서 다른 PC 사용자의 Downloads 경로가 config에 박혀
    있어 결과 기록이 조용히 빠진 일이 있었다(2026-08-18 확인). 없으면 저장소
    상위로 올라가며 `VXvue/` 루트에서 찾는다 — `VXvue/CLAUDE.md` 4절이 이
    파일을 **지식파일 폴더가 아니라 VXvue 루트**에 둔다고 명시한다.
    """
    override = (cfg.get("checklist_xlsx") or "").strip()
    if override and os.path.isfile(override):
        return override
    here = os.path.abspath(root or os.path.dirname(os.path.dirname(
        os.path.abspath(__file__))))
    for _ in range(4):                      # auto -> VXvue -> 자동화 ...
        here = os.path.dirname(here)
        if not here:
            break
        for candidate in (os.path.join(here, CHECKLIST_NAME),
                          os.path.join(here, "VXvue", CHECKLIST_NAME)):
            if os.path.isfile(candidate):
                return candidate
    return ""


def _find_header_row(ws, tc_col_name=TC_COL_HEADER):
    for row in range(1, min(ws.max_row, 20) + 1):
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row, col).value or "").strip() == tc_col_name:
                return row, col
    raise ValueError("'%s' 헤더를 찾지 못했습니다." % tc_col_name)


def _pick_tc_sheet(wb):
    for ws in wb:
        for row in range(1, min(ws.max_row, 20) + 1):
            for col in range(1, ws.max_column + 1):
                if str(ws.cell(row, col).value or "").strip() == TC_COL_HEADER:
                    return ws
    return wb.active


def write_results(source_xlsx, results, out_path=None, sheet_name=None, env=None):
    """판정 결과를 체크리스트 사본에 기록하고 결과 요약을 반환한다.

    results: TCResult 리스트. TC ID로 행을 매칭한다.
    env: `core/result.collect_env()` 결과. 있으면 1~5행에 실측값을 나란히 적는다.

    반환: {"path", "written", "extra", "sheet", "not_run"}
    """
    if not source_xlsx or not os.path.isfile(source_xlsx):
        raise FileNotFoundError(source_xlsx or "(체크리스트 경로를 찾지 못했습니다)")

    out_path = out_path or os.path.join(
        os.path.dirname(source_xlsx) or ".",
        "Checklist_Result_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M%S"))
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    shutil.copyfile(source_xlsx, out_path)

    wb = load_workbook(out_path)
    ws = wb[sheet_name] if sheet_name else (
        wb[CHECKLIST_SHEET] if CHECKLIST_SHEET in wb.sheetnames else _pick_tc_sheet(wb))
    hdr_row, tc_col = _find_header_row(ws)

    existing = dict((str(ws.cell(hdr_row, c).value or "").strip(), c)
                    for c in range(1, ws.max_column + 1))
    first_new = ws.max_column + 1
    col_of = {}
    for i, name in enumerate(RESULT_HEADERS):
        col_of[name] = existing.get(name, first_new + i)
        cell = ws.cell(hdr_row, col_of[name], name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 환경 실측값: 기존 K열 값을 덮어쓰지 않고 새 열에 나란히 적는다.
    if env:
        env_col = col_of[RESULT_HEADERS[0]]
        ws.cell(hdr_row - 1 if hdr_row > 1 else hdr_row, env_col,
                "자동화 실측(환경)").font = Font(bold=True, italic=True)
        for row in range(1, hdr_row):
            label = str(ws.cell(row, 1).value or "").strip()
            mapping = ENV_ROW_LABELS.get(label)
            if not mapping:
                continue
            section, key = mapping
            measured = (env.get(section) or {}).get(key)
            if measured in (None, ""):
                continue
            c = ws.cell(row, env_col, str(measured))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.font = Font(italic=True)

    by_id = {}
    for r in results:
        by_id.setdefault(r.tc_id, []).append(r)

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    written, not_run, matched_ids = 0, 0, set()

    for row in range(hdr_row + 1, ws.max_row + 1):
        tc_id = str(ws.cell(row, tc_col).value or "").strip()
        if not tc_id:
            continue
        hits = by_id.get(tc_id)
        if not hits:
            _set(ws, row, col_of["자동화 판정"], NOT_RUN)
            _set(ws, row, col_of["판정 일시"], stamp)
            ws.cell(row, col_of["수동 확인 항목"],
                    "이번 실행에 포함되지 않았다(자동화 코드 없음 또는 실행 범위 밖). "
                    "빈칸이 아니라 '미수행'으로 남긴다 — 확인했는데 이상 없음으로 "
                    "오해되지 않게 하기 위함.")
            not_run += 1
            continue
        matched_ids.add(tc_id)
        _write_row(ws, row, col_of, hits, stamp)
        written += 1

    # 체크리스트에 없는 TC ID(Precondition, VXvue_License, TC_Setting_ExportImport 등)
    extra = [tid for tid in by_id if tid not in matched_ids]
    if extra:
        row = ws.max_row + 2
        ws.cell(row, tc_col, "자동화 추가 항목").font = Font(bold=True)
        ws.cell(row, tc_col + 1,
                "체크리스트 원본에 대응 TC ID가 없는 자동화 항목"
                "(선행조건 점검·라이선스 확인 등)").font = Font(italic=True)
        for tid in sorted(extra):
            row += 1
            ws.cell(row, tc_col, tid)
            ws.cell(row, tc_col + 1, by_id[tid][0].title)
            _write_row(ws, row, col_of, by_id[tid], stamp)

    for name in RESULT_HEADERS:
        letter = ws.cell(hdr_row, col_of[name]).column_letter
        ws.column_dimensions[letter].width = (
            12 if name in ("자동화 판정", "PASS", "FAIL", "MANUAL", "SKIP", "BLOCKED")
            else 40)

    wb.save(out_path)
    return {"path": out_path, "written": written, "extra": len(extra),
            "sheet": ws.title, "not_run": not_run}


def _set(ws, row, col, value, status=None):
    cell = ws.cell(row, col, value)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    key = status or (value if value in FILLS else None)
    if key in FILLS:
        cell.fill = FILLS[key]
        cell.font = FONTS[key]
    return cell


def _write_row(ws, row, col_of, hits, stamp):
    verdicts = [h.verdict for h in hits]
    verdict = (FAIL if FAIL in verdicts else
               BLOCKED if BLOCKED in verdicts else
               MANUAL if MANUAL in verdicts else
               PASS if PASS in verdicts else SKIP)

    total = dict((k, 0) for k in (PASS, FAIL, MANUAL, SKIP, BLOCKED))
    fails, manuals, evidence = [], [], []
    for h in hits:
        for k, v in h.counts.items():
            total[k] = total.get(k, 0) + v
        for c in h.checks:
            if c.status == FAIL:
                fails.append("[Step %s] %s — 기대=%s / 실제=%s"
                             % (c.step, c.title, c.expected, c.actual))
            elif c.status in (MANUAL, BLOCKED):
                manuals.append("[%s Step %s] %s%s"
                               % (c.status, c.step, c.title,
                                  " — %s" % c.note if c.note else ""))
        evidence.extend(h.evidence)

    _set(ws, row, col_of["자동화 판정"], verdict, verdict)
    _set(ws, row, col_of["판정 일시"], stamp)
    for k in (PASS, FAIL, MANUAL, SKIP, BLOCKED):
        _set(ws, row, col_of[k], total[k])

    for name, items in (("실패 항목", fails), ("수동 확인 항목", manuals),
                        ("증적", evidence)):
        cell = ws.cell(row, col_of[name], "\n".join(items) if items else "")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
